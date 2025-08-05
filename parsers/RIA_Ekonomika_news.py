from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import os
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles import Font
import time

def extract_news():
    """
    Функция для извлечения 100 уникальных новостей с сайта RIA.ru/economy/
    """
    options = webdriver.ChromeOptions()
    options.add_argument("--headless")
    driver = webdriver.Chrome(options=options)

    try:
        print("[INFO] Открываем страницу https://ria.ru/economy/")
        driver.get("https://ria.ru/economy/")
        time.sleep(1)

        # Принимаем cookie, если есть
        try:
            cookie_btn = driver.find_element(By.CLASS_NAME, "cookie-warning__accept")
            cookie_btn.click()
            print("[INFO] Приняли cookie-уведомление")
            time.sleep(1)
        except:
            print("[INFO] Cookie-уведомление не найдено")

        news_data = {'name': [], 'link': [], 'date': []}
        seen_links = set()
        max_news = 100

        print("[INFO] Прокручиваем страницу...")
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight - 300);")
        time.sleep(1)

        try:
            load_more_button = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.CLASS_NAME, "list-more"))
            )
            driver.execute_script("window.scrollBy(0, 200);")
            time.sleep(0.5)
            load_more_button.click()
            print("[INFO] Нажата кнопка 'Еще материалы'")
            time.sleep(1)
        except Exception as e:
            print(f"[WARN] Не удалось нажать на кнопку 'Еще материалы': {str(e)}")

        print("[INFO] Начинаем сбор новостей...")

        for i in range(5):
            print(f"[INFO] Прокрутка {i+1}/5")
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight - 200);")
            time.sleep(1)

            try:
                load_more_button = driver.find_element(By.CLASS_NAME, "list-more")
                driver.execute_script("arguments[0].scrollIntoView();", load_more_button)
                driver.execute_script("window.scrollBy(0, -100);")
                load_more_button.click()
                print("[INFO] Нажата кнопка 'Еще материалы'")
                time.sleep(1)
            except:
                print("[INFO] Кнопка 'Еще материалы' не найдена")

            # Обновляем список новостей после каждой прокрутки
            news_items = driver.find_elements(By.CSS_SELECTOR, "div.list-item")

            for item in news_items:
                if len(news_data['name']) >= max_news:
                    break

                try:
                    title_element = item.find_element(By.CSS_SELECTOR, "a.list-item__title")
                    title = title_element.text.strip()
                    link = title_element.get_attribute("href")

                    if not link or link in seen_links:
                        continue

                    seen_links.add(link)

                    time_element = item.find_element(By.CSS_SELECTOR, "div.list-item__info-item[data-type='date']")
                    time_text = time_element.text.strip()

                    news_data['name'].append(title)
                    news_data['link'].append(link)
                    news_data['date'].append(time_text)

                    print(f"[{len(news_data['name'])}/{max_news}] Собрана новость: {title[:50]}... | Время: {time_text}")

                    # 👉 Прогресс для UI
                    progress_percentage = int((len(news_data['name']) / max_news) * 70)
                    print(f"Progress: {progress_percentage}% [{len(news_data['name'])}/{max_news}]")

                except Exception as e:
                    print(f"[WARN] Ошибка при обработке новости: {str(e)}")
                    continue

        print(f"[INFO] Завершён сбор. Собрано уникальных новостей: {len(news_data['name'])}")

    except Exception as e:
        print(f"[ERROR] Критическая ошибка: {str(e)}")
        raise
    finally:
        driver.quit()
        print("[INFO] Браузер закрыт.")

    return news_data


def save_to_excel(news_info, output_file_name="News_data_RIA_First_100_News.xlsx"):
    """
    Сохраняет новости в Excel с форматированием и гиперссылками.
    """
    print("Progress: 98% [Сохраняем в Excel]")

    # Создаём папку для сохранения Excel-файлов
    save_dir = os.path.join(os.getcwd(), "parsed_excels")
    os.makedirs(save_dir, exist_ok=True)

    # Полный путь к выходному файлу
    full_output_path = os.path.join(save_dir, output_file_name)

    print(f"[INFO] Сохраняем данные в файл {full_output_path}")

    try:
        data_for_df = {
            'Название': news_info.get('name', []),
            'Ссылка': news_info.get('link', []),
            'Время публикации': news_info.get('date', [])
        }

        if not any(data_for_df.values()):
            raise ValueError("Нет данных для сохранения")

        df = pd.DataFrame(data_for_df)

        with pd.ExcelWriter(full_output_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Новости')

            workbook = writer.book
            worksheet = writer.sheets['Новости']

            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[column_letter].width = adjusted_width

            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=2)
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("http"):
                    cell.hyperlink = Hyperlink(ref=cell.coordinate, target=cell.value, tooltip="Перейти по ссылке")
                    cell.font = Font(color="0563C1", underline="single")

        print(f"[SUCCESS] Данные успешно сохранены в '{full_output_path}'")
        print("Progress: 100% [Готово]")

    except Exception as e:
        print(f"[ERROR] Ошибка при сохранении в Excel: {str(e)}")
        raise

if __name__ == "__main__":
    try:
        print("=== НАЧАЛО РАБОТЫ СКРИПТА ===")
        print("\n[1/2] Сбор новостей...")
        news_info = extract_news()

        print("\n[2/2] Сохранение результатов...")
        save_to_excel(news_info)

        print("\n=== СКРИПТ УСПЕШНО ЗАВЕРШЕН ===")

    except Exception as e:
        print(f"\n=== ОШИБКА: {str(e)} ===")
