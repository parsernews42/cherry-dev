from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import os
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles import Font
import time
from tqdm import tqdm  # Импортируем tqdm для отображения прогресса

def extract_news():
    """
    Функция для извлечения 50 новостей с сайта 1prime.ru/state_regulation/
    с оптимизированным скроллингом для поиска кнопки "Ещё"
    """
    # Настройка веб-драйвера
    options = webdriver.ChromeOptions()
    options.add_argument("--headless")  # Запуск в фоновом режиме
    driver = webdriver.Chrome(options=options)

    try:
        print("[INFO] Открываем страницу https://1prime.ru/state_regulation/")
        driver.get("https://1prime.ru/state_regulation/")
        time.sleep(2)  # Увеличили время ожидания первичной загрузки

        # Инициализация данных
        news_data = {'name': [], 'link': [], 'date': []}
        max_news = 50  # Целевое количество новостей
        scroll_attempts = 0
        max_scroll_attempts = 10  # Максимальное количество попыток скроллинга

        print("[INFO] Начинаем сбор новостей...")

        while len(news_data['name']) < max_news and scroll_attempts < max_scroll_attempts:
            # Собираем текущие новости перед скроллингом
            news_items = driver.find_elements(By.CSS_SELECTOR, "div.list-item")
            print(f"[DEBUG] Всего найдено новостей: {len(news_items)}")

            # Парсинг новостных блоков с использованием tqdm для отображения прогресса
            for item in tqdm(news_items[len(news_data['name']):], desc="Сбор новостей", total=max_news - len(news_data['name'])):
                try:
                    # Извлекаем заголовок и ссылку
                    title_element = item.find_element(By.CSS_SELECTOR, "a.list-item__title")
                    title = title_element.text.strip()
                    link = title_element.get_attribute("href")

                    # Извлекаем время публикации
                    time_element = item.find_element(By.CSS_SELECTOR, "div.list-item__info div.list-item__date")
                    time_text = time_element.text.strip()

                    # Заполняем данные
                    news_data['name'].append(title)
                    news_data['link'].append(link)
                    news_data['date'].append(time_text)

                    print(f"[{len(news_data['name'])}/{max_news}] Собрана новость: {title[:50]}... | Время: {time_text}")

                    if len(news_data['name']) >= max_news:
                        break

                except Exception as e:
                    print(f"[WARN] Ошибка при обработке новости: {str(e)}")
                    continue

            if len(news_data['name']) >= max_news:
                break

            # Плавный скроллинг вниз (эквивалент 10 щелчков колесика мыши)
            print(f"[INFO] Скроллинг вниз (попытка {scroll_attempts + 1}/{max_scroll_attempts})")
            driver.execute_script("window.scrollBy(0, 1000);")  # Скролл на ~10 щелчков
            time.sleep(1.5)  # Даем время для загрузки контента

            # Пробуем найти и нажать кнопку "Ещё материалы"
            try:
                load_more_button = WebDriverWait(driver, 3).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, "div.list-more"))
                )
                # Прокручиваем немного вверх, чтобы кнопка была видна
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", load_more_button)
                time.sleep(0.5)
                load_more_button.click()
                print("[INFO] Нажата кнопка 'Еще материалы'")
                time.sleep(2)  # Даем время для загрузки новых новостей
            except Exception as e:
                print(f"[INFO] Кнопка 'Еще материалы' не найдена или не кликабельна: {str(e)}")

            scroll_attempts += 1

        print(f"[INFO] Завершён сбор новостей. Собрано {len(news_data['name'])} новостей.")

    except Exception as e:
        print(f"[ERROR] Критическая ошибка: {str(e)}")
        raise
    finally:
        driver.quit()
        print("[INFO] Браузер закрыт.")

    return news_data

def save_to_excel(news_info, output_file_name="News_data_1prime_First_50_News.xlsx"):
    """
    Сохраняет данные новостей в файл Excel с автошириной столбцов,
    задает заголовки, и вставляет ссылки как гиперссылки.
    """
    # Создаём папку для сохранения Excel-файлов
    save_dir = os.path.join(os.getcwd(), "parsed_excels")
    os.makedirs(save_dir, exist_ok=True)

    # Полный путь к выходному файлу
    full_output_path = os.path.join(save_dir, output_file_name)

    print(f"[INFO] Сохраняем данные в файл {full_output_path}")

    try:
        # Создаем DataFrame
        data_for_df = {
            'Название': news_info.get('name', []),
            'Ссылка': news_info.get('link', []),
            'Время публикации': news_info.get('date', [])
        }

        if not any(data_for_df.values()):
            raise ValueError("Нет данных для сохранения")

        df = pd.DataFrame(data_for_df)

        # Сохраняем в Excel
        with pd.ExcelWriter(full_output_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Новости')

            # Получаем книгу и лист для дополнительного форматирования
            workbook = writer.book
            worksheet = writer.sheets['Новости']

            # Настраиваем ширину столбцов
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)

                # Определяем максимальную длину содержимого в столбце
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                # Устанавливаем ширину с небольшим запасом
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # Форматируем столбец с ссылками
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=2)  # Столбец B (ссылки)
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("http"):
                    cell.hyperlink = Hyperlink(ref=cell.coordinate, target=cell.value,
                                              tooltip="Перейти по ссылке")
                    cell.font = Font(color="0563C1", underline="single")

        print(f"[SUCCESS] Данные успешно сохранены в файл '{full_output_path}'")

    except Exception as e:
        print(f"[ERROR] Ошибка при сохранении в Excel: {str(e)}")
        raise

if __name__ == "__main__":
    try:
        print("=== НАЧАЛО РАБОТЫ СКРИПТА ===")

        # Сбор новостей
        print("\n[1/2] Запускаем сбор новостей...")
        news_info = extract_news()

        # Сохранение результатов
        print("\n[2/2] Сохраняем результаты...")
        save_to_excel(news_info)

        print("\n=== СКРИПТ УСПЕШНО ЗАВЕРШЕН ===")

    except Exception as e:
        print(f"\n=== ОШИБКА ВЫПОЛНЕНИЯ: {str(e)} ===")
