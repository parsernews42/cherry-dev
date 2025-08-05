from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import os
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles import Font
import time


def extract_news():
    """
    Функция для извлечения до 100 новостей с сайта mashnews.ru с названиями, ссылками и датами публикаций.
    Поиск ограничен 3 прокрутками вниз без прокрутки вверх.
    """
    # Настройка веб-драйвера
    options = webdriver.ChromeOptions()
    options.add_argument("--headless")  # Запуск в фоновом режиме
    driver = webdriver.Chrome(options=options)  # Убедитесь, что версия ChromeDriver соответствует версии Chrome

    try:
        print("[INFO] Открываем страницу https://mashnews.ru/publications/")
        driver.get("https://mashnews.ru/publications/")

        print("[INFO] Ждем загрузки элемента с новостями")
        WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.ID, "thunder")))

        news_data = {'name': [], 'link': [], 'date': []}
        processed_articles = set()
        max_scrolls = 3
        scrolls_count = 0
        max_news = 50  # Изменить эту переменную, если нужно больше или меньше новостей

        while scrolls_count < max_scrolls and len(news_data['name']) < max_news:
            scrolls_count += 1
            print(f"[INFO] Выполняем прокрутку вниз {scrolls_count}/{max_scrolls}")

            scroll_height = driver.execute_script("return document.body.scrollHeight")
            scroll_step = scroll_height / max_scrolls
            target_scroll = scroll_step * scrolls_count
            driver.execute_script(f"window.scrollTo(0, {target_scroll});")

            time.sleep(5)  # Ожидание загрузки

            all_articles = driver.find_elements(By.CSS_SELECTOR, "#thunder > div > div")
            print(f"[DEBUG] Найдено элементов на странице: {len(all_articles)}")

            for article in all_articles:
                try:
                    article_id = article.id
                except:
                    article_id = None
                if article_id in processed_articles:
                    continue

                # Попытка получить дату публикации (если есть)
                try:
                    date_elem = article.find_element(By.CLASS_NAME, "thunder-month")
                    current_section_date = date_elem.text.strip()
                except Exception:
                    current_section_date = ""

                # Попытка получить время публикации (если есть)
                try:
                    time_elem = article.find_element(By.CLASS_NAME, "thunder-time")
                    time_text = time_elem.text.strip()
                except Exception:
                    time_text = ""

                # Полная дата-метка
                full_date_str = (current_section_date + " " + time_text).strip()

                # Ищем название и ссылку
                try:
                    link_element = article.find_element(By.CLASS_NAME, "thunder-link")
                    link = link_element.get_attribute("href")
                    try:
                        title_element = link_element.find_element(By.TAG_NAME, "strong")
                        title = title_element.text.strip()
                    except Exception:
                        title = link_element.text.strip()
                except Exception as e:
                    print(f"[WARN] Не удалось получить название или ссылку новости: {e}")
                    continue

                if not link:
                    link = "Ссылка не найдена"

                try:
                    print(f"[DEBUG] Новость: '{title}', Ссылка: {link}, Дата и время: {full_date_str}")
                except UnicodeEncodeError:
                    print(f"[DEBUG] Новость (неотображаемые символы): {title.encode('ascii', errors='replace')}")

                news_data['name'].append(title)
                news_data['link'].append(link)
                news_data['date'].append(full_date_str)

                processed_articles.add(article_id)

                # Обновляем прогресс
                progress_percentage = int((len(news_data['name']) / max_news) * 100)
                print(f"\rProgress: {progress_percentage}% [{len(news_data['name'])}/{max_news}]", end="", flush=True)  # Обновление прогресса

                if len(news_data['name']) >= max_news:
                    print(f"\n[INFO] Достигнуто {max_news} новостей, прекращаем сбор.")
                    break

            if len(news_data['name']) >= max_news:
                break

        print(f"\n[INFO] Завершён сбор новостей. Собрано {len(news_data['name'])} новостей.")

    finally:
        driver.quit()
        print("[INFO] Закрыт браузер.")

    return news_data

def save_to_excel(news_info, output_file_name):
    """
    Сохраняет данные новостей в файл Excel с автошириной столбцов,
    задает заголовки, и вставляет ссылки как гиперссылки.
    """
    # Создаём папку для сохранения Excel-файлов
    save_dir = os.path.join(os.getcwd(), "parsed_excels")
    os.makedirs(save_dir, exist_ok=True)

    # Полный путь к выходному файлу
    full_output_path = os.path.join(save_dir, output_file_name)

    print(f"[INFO] Сохраняем данные в файл {full_output_path}")
    data_for_df = {
        'Название': news_info.get('name', []),
        'Ссылка': news_info.get('link', []),
        'Дата публикации': news_info.get('date', [])
    }
    df = pd.DataFrame(data_for_df)

    df.to_excel(full_output_path, index=False, sheet_name='Новости')

    # Открываем книгу Excel
    wb = load_workbook(full_output_path)
    ws = wb['Новости']

    # Заполняем данные в Excel
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for cell in column_cells:
            if cell.value:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        ws.column_dimensions[column_letter].width = max_length + 2

    link_col = 2
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=link_col)
        link = cell.value
        if link and isinstance(link, str) and link.startswith("http"):
            # Создаем объект Hyperlink
            hyperlink = Hyperlink(ref=cell.coordinate, target=link, tooltip="Перейти по ссылке")
            cell.hyperlink = hyperlink
            cell.style = "Hyperlink"
            # Опционально: установить синий цвет и подчеркнутый шрифт вручную (обычно стиль Hyperlink это делает)
            cell.font = Font(color="0000EE", underline="single")

    wb.save(full_output_path)
    print(f"[INFO] Данные успешно сохранены и отформатированы в файле '{full_output_path}'.")

if __name__ == "__main__":
    # Сбор новостей с сайта
    news_info = extract_news()

    # Сохранение данных в файл Excel
    file_name = "News_data_Mashnews_First_50_News.xlsx"
    save_to_excel(news_info, file_name)
    print("[INFO] Скрипт завершен успешно.")
