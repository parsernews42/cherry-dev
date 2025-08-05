from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import os
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles import Font
import time

def extract_news(progress_callback):
    """
    Функция для извлечения новостей с сайта https://www.interfax-russia.ru/main?per-page=100
    Собирает название и ссылку на новость.
    """
    # Настройка веб-драйвера
    options = webdriver.ChromeOptions()
    options.add_argument("--headless")  # Запуск в фоновом режиме
    driver = webdriver.Chrome(options=options)  # Убедитесь, что версия ChromeDriver соответствует версии Chrome

    try:
        url = "https://www.interfax-russia.ru/main?per-page=100"
        print(f"[INFO] Открываем страницу {url}")
        driver.get(url)

        # Ждем загрузки блока новостей
        news_container_xpath = "//div[@class='col-12 col-xl-8 mt-0']//ul"
        print("[INFO] Ждем загрузки блока новостей")
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.XPATH, news_container_xpath))
        )

        news_data = {'name': [], 'link': []}

        # Получаем все новости
        news_items = driver.find_elements(By.XPATH, news_container_xpath + "/li")
        print(f"[INFO] Найдено новостей на странице: {len(news_items)}")

        total_items = len(news_items)

        for idx, item in enumerate(news_items):
            try:
                # Название новости из атрибута alt изображения
                title_elem = item.find_element(By.XPATH, ".//img[@class='img-fluid w-100']")
                title = title_elem.get_attribute("alt")
                if title is not None:
                    title = title.strip()  # Применяем strip() только если title не None
                # Ссылка новости
                link_elem = item.find_element(By.XPATH, ".//a[@class='stretched-link']")
                link = link_elem.get_attribute("href")

                print(f"[DEBUG] Новость: '{title}', Ссылка: {link}")

                news_data['name'].append(title)
                news_data['link'].append(link)

                # Обновляем прогресс
                progress = int((idx + 1) / total_items * 100)
                progress_callback(progress)

            except Exception as e:
                print(f"[WARN] Ошибка при извлечении названия или ссылки: {e}")
                continue

        print(f"[INFO] Завершён сбор новостей. Собрано {len(news_data['name'])} новостей.")

    finally:
        driver.quit()
        print("[INFO] Закрыт браузер.")

    return news_data

def save_to_excel(news_info, output_file_name):
    # Создаём папку для сохранения Excel-файлов
    save_dir = os.path.join(os.getcwd(), "parsed_excels")
    os.makedirs(save_dir, exist_ok=True)

    # Полный путь к выходному файлу
    full_output_path = os.path.join(save_dir, output_file_name)

    print(f"[INFO] Сохраняем данные в файл {full_output_path}")
    data_for_df = {
        'Название': news_info.get('name', []),
        'Ссылка': news_info.get('link', [])
    }
    df = pd.DataFrame(data_for_df)

    df.to_excel(full_output_path, index=False, sheet_name='Новости')

    wb = load_workbook(full_output_path)
    ws = wb['Новости']

    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for cell in column_cells:
            if cell.value:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        ws.column_dimensions[column_letter].width = max_length + 2

    link_col = 2
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=link_col)
        link = cell.value
        if link and isinstance(link, str) and link.startswith("http"):
            hyperlink = Hyperlink(ref=cell.coordinate, target=link, tooltip="Перейти по ссылке")
            cell.hyperlink = hyperlink
            cell.style = "Hyperlink"
            cell.font = Font(color="0000EE", underline="single")

    wb.save(full_output_path)
    print(f"[INFO] Данные успешно сохранены и отформатированы в файле '{full_output_path}'.")

if __name__ == "__main__":
    # Пример использования функции с прогрессом
    def progress_callback(progress):
        print(f"Progress: {progress}%")  # Здесь можно заменить на сигнал для UI

    news_info = extract_news(progress_callback)
    file_name = "News_data_Interfax_100_News.xlsx"
    save_to_excel(news_info, file_name)
    print("[INFO] Скрипт завершен успешно.")
