from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import os
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles import Font
from datetime import datetime, timedelta
import time

def extract_news(progress_callback):
    """
    Извлечение новостей с https://www.interfax.ru/business/
    Только за сегодня и вчера, прогресс только по сохраняемым новостям
    """
    options = webdriver.ChromeOptions()
    options.add_argument("--headless")
    driver = webdriver.Chrome(options=options)

    try:
        url = "https://www.interfax.ru/business/"
        print(f"[INFO] Открываем страницу {url}")
        driver.get(url)

        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "div.timeline"))
        )

        # Загрузка дополнительных новостей (максимум 3 клика)
        for i in range(3):
            try:
                news_blocks = driver.find_elements(
                    By.XPATH,
                    '//div[contains(@class, "timeline")]//div[contains(@class, "timeline__group")]//div[not(contains(@class, "timeline__more"))]'
                )
                current_count = len(news_blocks)

                more_button = driver.find_element(By.CSS_SELECTOR, "div.timeline__more")
                driver.execute_script("arguments[0].click();", more_button)
                print(f"[INFO] Клик по 'Загрузить еще новости' ({i+1}/3)")

                WebDriverWait(driver, 10).until(
                    lambda d: len(d.find_elements(
                        By.XPATH,
                        '//div[contains(@class, "timeline")]//div[contains(@class, "timeline__group")]//div[not(contains(@class, "timeline__more"))]'
                    )) > current_count
                )
                time.sleep(0.5)

            except Exception:
                print("[INFO] Больше нет кнопки или новые новости не загрузились.")
                break

        all_news_blocks = driver.find_elements(
            By.XPATH,
            '//div[contains(@class, "timeline")]//div[contains(@class, "timeline__group")]//div[not(contains(@class, "timeline__more"))]'
        )
        print(f"[INFO] Всего найдено блоков новостей: {len(all_news_blocks)}")

        # Подготовка фильтра по дате
        today = datetime.now()
        yesterday = today - timedelta(days=1)
        allowed_dates = {
            today.strftime("%d-%m-%Y"),
            yesterday.strftime("%d-%m-%Y")
        }

        # Предварительный сбор только актуальных блоков
        relevant_blocks = []
        for item in all_news_blocks:
            try:
                time_elem = item.find_element(By.TAG_NAME, "time")
                datetime_str = time_elem.get_attribute("datetime")
                if not datetime_str:
                    continue
                dt_obj = datetime.strptime(datetime_str, "%Y-%m-%dT%H:%M")
                if dt_obj.strftime("%d-%m-%Y") in allowed_dates:
                    relevant_blocks.append((item, dt_obj))
            except:
                continue

        print(f"[INFO] Отобрано актуальных новостей (сегодня и вчера): {len(relevant_blocks)}")

        news_data = {'name': [], 'link': [], 'date': []}
        total_relevant = len(relevant_blocks)

        for idx, (item, dt_obj) in enumerate(relevant_blocks):
            try:
                a_tag = item.find_element(By.TAG_NAME, "a")
                link = a_tag.get_attribute("href")
                title = a_tag.find_element(By.TAG_NAME, "h3").text.strip()

                formatted_date = dt_obj.strftime("%d-%m-%Y %H:%M")

                news_data['name'].append(title)
                news_data['link'].append(link)
                news_data['date'].append(formatted_date)

                print(f"[DEBUG] {idx+1}/{total_relevant}: '{title}' | {formatted_date} | {link}")
                progress = int((idx + 1) / total_relevant * 100)
                progress_callback(progress)

            except Exception as e:
                print(f"[WARN] Ошибка при обработке новости #{idx+1}: {e}")
                continue

    finally:
        driver.quit()
        print("[INFO] Закрыт браузер.")

    return news_data


def save_to_excel(news_info, output_file_name):
    # Создаём папку для сохранения Excel-файлов
    save_dir = os.path.join(os.getcwd(), "parsed_excels")
    os.makedirs(save_dir, exist_ok=True)

    # Полный путь к выходному файлу
    full_output_path = os.path.join(save_dir, output_file_name)

    print(f"[INFO] Сохраняем данные в файл {full_output_path}")
    data_for_df = {
        'Название': news_info.get('name', []),
        'Ссылка': news_info.get('link', []),
        'Дата': news_info.get('date', [])
    }
    df = pd.DataFrame(data_for_df)

    df.to_excel(full_output_path, index=False, sheet_name='Новости')

    wb = load_workbook(full_output_path)
    ws = wb['Новости']

    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for cell in column_cells:
            if cell.value:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        ws.column_dimensions[column_letter].width = max_length + 2

    link_col = 2
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=link_col)
        link = cell.value
        if link and isinstance(link, str) and link.startswith("http"):
            hyperlink = Hyperlink(ref=cell.coordinate, target=link, tooltip="Перейти по ссылке")
            cell.hyperlink = hyperlink
            cell.style = "Hyperlink"
            cell.font = Font(color="0000EE", underline="single")

    wb.save(full_output_path)
    print(f"[INFO] Данные успешно сохранены в файле '{full_output_path}'.")


if __name__ == "__main__":
    def progress_callback(progress):
        print(f"Progress: {progress}%")

    news_info = extract_news(progress_callback)
    file_name = "News_data_Interfax_Today_Yesterday.xlsx"
    save_to_excel(news_info, file_name)
    print("[INFO] Скрипт завершен успешно.")
