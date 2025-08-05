from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import os
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles import Font
import time

def extract_news():
    """
    Функция для извлечения до 300 новостей с tass.ru (экономика)
    """
    options = webdriver.ChromeOptions()
    #options.add_argument("--headless")
    driver = webdriver.Chrome(options=options)
    driver.maximize_window()

    try:
        driver.get("https://tass.ru/ekonomika")

        # Ждём появления первых новостей
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.CLASS_NAME, "tass_pkg_link-v5WdK"))
        )

        print("Прокручиваем страницу вниз для появления кнопки...")
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(2)

        # Пробуем нажать на кнопку "Загрузить больше результатов"
        try:
            load_more_button = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, '//*[@id="infinite_listing"]/button'))
            )
            driver.execute_script("arguments[0].scrollIntoView();", load_more_button)
            time.sleep(1)
            load_more_button.click()
            print("Кнопка 'Загрузить больше результатов' нажата.")
            time.sleep(3)
        except Exception as e:
            print("Кнопка не нажалась или не найдена:", str(e))

        # Продолжаем прокручивать страницу до 300 новостей
        max_scrolls = 50
        for scroll in range(max_scrolls):
            articles = driver.find_elements(By.CLASS_NAME, "tass_pkg_link-v5WdK")
            print(f"Сейчас загружено новостей: {len(articles)}")

            if len(articles) >= 300:
                print("Достигнуто 300 новостей.")
                break

            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(2.5)

        # Сбор новостей
        articles = driver.find_elements(By.CLASS_NAME, "tass_pkg_link-v5WdK")
        total_to_collect = min(300, len(articles))
        print(f"Переходим к сбору {total_to_collect} новостей...")

        news_data = {'name': [], 'link': [], 'date': []}

        for idx, article in enumerate(articles[:total_to_collect], start=1):
            try:
                title = article.find_element(By.CLASS_NAME, "tass_pkg_title-xVUT1").text
                link = article.get_attribute("href") or "Ссылка не найдена"

                try:
                    date_element = article.find_element(By.CLASS_NAME, "tass_pkg_marker-JPOGl")
                    date = date_element.text if date_element else "Дата не найдена"
                except:
                    date = "Дата не найдена"

                news_data['name'].append(title)
                news_data['link'].append(link)
                news_data['date'].append(date)

                progress_percentage = int((idx / total_to_collect) * 100)
                print(f"\rОбработка: {progress_percentage}% [{idx}/{total_to_collect}]", end="", flush=True)

            except Exception as e:
                print(f"\nОшибка при обработке новости #{idx}: {str(e)}")
                continue

        print("\nСбор завершён.")
        return news_data

    finally:
        driver.quit()

def save_to_excel(news_info, output_file_name):
    """
    Сохраняет данные в Excel файл с форматированием
    """
    # Создаём папку для сохранения Excel-файлов
    save_dir = os.path.join(os.getcwd(), "parsed_excels")
    os.makedirs(save_dir, exist_ok=True)

    # Полный путь к выходному файлу
    full_output_path = os.path.join(save_dir, output_file_name)

    print(f"[INFO] Сохраняем данные в файл {full_output_path}")
    df = pd.DataFrame({
        'Название': news_info.get('name', []),
        'Ссылка': news_info.get('link', []),
        'Дата публикации': news_info.get('date', [])
    })

    df.to_excel(full_output_path, index=False, sheet_name='Новости')

    wb = load_workbook(full_output_path)
    ws = wb['Новости']

    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value is not None)
        if col[0].column is not None:
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=2)
        if cell.value and isinstance(cell.value, str) and cell.value.startswith("http"):
            cell.hyperlink = Hyperlink(ref=cell.coordinate, target=cell.value)
            cell.font = Font(color="0000EE", underline="single")

    wb.save(full_output_path)
    print(f"Данные сохранены в файл: {full_output_path}")

if __name__ == '__main__':
    print("=== Парсер новостей TASS ===")
    news_data = extract_news()
    save_to_excel(news_data, 'News_data_Tass_Ekonomika.xlsx')
    print("Работа завершена успешно!")
