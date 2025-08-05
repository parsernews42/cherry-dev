from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import os
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles import Font
import time

def extract_news():
    """
    Функция для извлечения новостей с сайта rg.ru в разделе экономики.
    """
    # Настройка веб-драйвера
    options = webdriver.ChromeOptions()
    options.add_argument("--headless")  # Запуск в фоновом режиме
    driver = webdriver.Chrome(options=options)  # Убедитесь, что версия ChromeDriver соответствует версии Chrome

    try:
        print("[INFO] Открываем страницу https://rg.ru/tema/ekonomika/business")
        driver.get("https://rg.ru/tema/ekonomika/business")

        print("[INFO] Ждем загрузки элемента с новостями")
        WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CLASS_NAME, "Page_main__CL9dG")))

        news_data = {'name': [], 'link': [], 'date': []}  # Словарь для хранения данных новостей
        max_scrolls = 3  # Максимальное количество прокруток
        scrolls_count = 0  # Счетчик прокруток
        max_news = 50  # Максимальное количество новостей для извлечения

        while scrolls_count < max_scrolls and len(news_data['name']) < max_news:
            scrolls_count += 1
            print(f"[INFO] Выполняем прокрутку вниз {scrolls_count}/{max_scrolls}")

            scroll_height = driver.execute_script("return document.body.scrollHeight")  # Получаем высоту страницы
            driver.execute_script(f"window.scrollTo(0, {scroll_height});")  # Прокручиваем страницу

            time.sleep(2)  # Ожидание загрузки

            all_articles = driver.find_elements(By.CLASS_NAME, "PageRubricContent_listItem__KVIae")  # Находим все статьи
            print(f"[DEBUG] Найдено элементов на странице: {len(all_articles)}")

            for article in all_articles:
                # Ищем название, ссылку и дату
                try:
                    link_element = article.find_element(By.TAG_NAME, "a")
                    link = link_element.get_attribute("href")  # Получаем ссылку
                    title_element = article.find_element(By.CLASS_NAME, "ItemOfListStandard_title__Ajjlf")
                    title = title_element.text.strip()  # Получаем название
                    date_element = article.find_element(By.CLASS_NAME, "ItemOfListStandard_datetime__GstJi")
                    date = date_element.text.strip()  # Получаем дату
                except Exception as e:
                    print(f"[WARN] Не удалось получить данные новости: {e}")
                    continue

                print(f"[DEBUG] Новость: '{title}', Ссылка: {link}, Дата: {date}")

                # Добавляем данные в словарь
                news_data['name'].append(title)
                news_data['link'].append(link)
                news_data['date'].append(date)

                # Обновляем прогресс
                progress_percentage = int((len(news_data['name']) / max_news) * 100)
                print(f"\rProgress: {progress_percentage}% [{len(news_data['name'])}/{max_news}]", end="", flush=True)  # Обновление прогресса

                if len(news_data['name']) >= max_news:  # Проверка на достижение максимума новостей
                    print(f"\n[INFO] Достигнуто {max_news} новостей, прекращаем сбор.")
                    break

            if len(news_data['name']) >= max_news:
                break

        print(f"\n[INFO] Завершён сбор новостей. Собрано {len(news_data['name'])} новостей.")

    finally:
        driver.quit()  # Закрываем браузер
        print("[INFO] Закрыт браузер.")

    return news_data  # Возвращаем собранные данные

def save_to_excel(news_info, output_file_name):
    """
    Сохраняет данные новостей в файл Excel с автошириной столбцов,
    задает заголовки, и вставляет ссылки как гиперссылки.
    """
    # Создаём папку для сохранения Excel-файлов
    save_dir = os.path.join(os.getcwd(), "parsed_excels")
    os.makedirs(save_dir, exist_ok=True)

    # Полный путь к выходному файлу
    full_output_path = os.path.join(save_dir, output_file_name)

    print(f"[INFO] Сохраняем данные в файл {full_output_path}")
    data_for_df = {
        'Название': news_info.get('name', []),
        'Ссылка': news_info.get('link', []),
        'Дата публикации': news_info.get('date', [])
    }
    df = pd.DataFrame(data_for_df)  # Создаем DataFrame из собранных данных

    df.to_excel(full_output_path, index=False, sheet_name='Новости')  # Сохраняем в Excel

    # Открываем книгу Excel
    wb = load_workbook(full_output_path)
    ws = wb['Новости']

    # Заполняем данные в Excel
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for cell in column_cells:
            if cell.value:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        ws.column_dimensions[column_letter].width = max_length + 2  # Устанавливаем ширину столбца

    link_col = 2  # Индекс столбца со ссылками
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=link_col)
        link = cell.value
        if link and isinstance(link, str) and link.startswith("http"):
            # Создаем объект Hyperlink
            hyperlink = Hyperlink(ref=cell.coordinate, target=link, tooltip="Перейти по ссылке")
            cell.hyperlink = hyperlink
            cell.style = "Hyperlink"  # Устанавливаем стиль гиперссылки
            # Опционально: установить синий цвет и подчеркнутый шрифт вручную
            cell.font = Font(color="0000EE", underline="single")

    wb.save(full_output_path)  # Сохраняем изменения в Excel
    print(f"[INFO] Данные успешно сохранены и отформатированы в файле '{full_output_path}'.")

if __name__ == "__main__":
    # Сбор новостей с сайта
    news_info = extract_news()

    # Сохранение данных в файл Excel
    file_name = "News_data_RG_First_50_News.xlsx"
    save_to_excel(news_info, file_name)
    print("[INFO] Скрипт завершен успешно.")
